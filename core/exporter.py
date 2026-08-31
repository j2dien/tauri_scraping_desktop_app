"""
exporter.py — Modul untuk export data ke file Excel (.xlsx).

Menghasilkan file Excel dengan format rapi dan terstruktur:
1. Mode Analisis Lengkap:
   - Summary: Ranking top commenters dan statistik total likes & komentar
   - Detail Komentar: Semua komentar lengkap, like komentar, like post, tanggal, dan caption
   - Daftar Postingan: Daftar postingan dengan jumlah likes, tanggal publish, dan caption
2. Mode Link Saja:
   - Daftar Link Postingan: URL, tanggal, tipe post, likes, dan caption
"""

import re
from pathlib import Path
from datetime import datetime
from typing import Any, List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Regex untuk membersihkan karakter kontrol ilegal di XML/openpyxl
ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff]")

BULAN_INDONESIA = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def clean_cell_value(val: Any) -> Any:
    """Membersihkan nilai sel dari karakter ilegal openpyxl."""
    if val is None:
        return ""
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val


def format_tanggal_indonesia(date_str: str) -> str:
    """
    Konversi format tanggal dari 'YYYY-MM-DD HH:MM:SS' ke format Indonesia
    'DD Bulan YYYY HH:MM:SS' (contoh: '20 Agustus 2026 09:08:49').
    """
    if not date_str or date_str == "N/A":
        return str(date_str or "N/A")

    try:
        # Jika format YYYY-MM-DD HH:MM:SS
        if " " in date_str and "-" in date_str:
            dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            bulan = BULAN_INDONESIA[dt.month]
            return f"{dt.day:02d} {bulan} {dt.year} {dt.strftime('%H:%M:%S')}"
        # Jika format DD-MM-YYYY
        elif "-" in date_str and len(date_str.split("-")[0]) == 2:
            dt = datetime.strptime(date_str, "%d-%m-%Y")
            bulan = BULAN_INDONESIA[dt.month]
            return f"{dt.day:02d} {bulan} {dt.year}"
        return str(date_str)
    except (ValueError, IndexError):
        return str(date_str)


def _style_header(ws, headers: list[str], fill_color: str = "1F4E79"):
    """Terapkan styling pada header row."""
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=clean_cell_value(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _auto_fit_columns(ws, min_width: int = 12, max_width: int = 55):
    """Auto-fit lebar kolom berdasarkan konten per baris (aman untuk multiline)."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0
        for cell in col:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_length = max(max_length, len(line))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def _save_workbook_safely(wb: Workbook, target_filename: str) -> str:
    """Simpan workbook dengan penanganan jika file sedang dibuka oleh user (PermissionError)."""
    target_path = Path(target_filename)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb.save(str(target_path))
        return str(target_path)
    except PermissionError:
        # Jika file sedang di-lock oleh Excel, tambahkan suffix waktu unik
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = target_path.with_name(f"{target_path.stem}_{timestamp}{target_path.suffix}")
        wb.save(str(alt_path))
        return str(alt_path)


def export_to_excel(
    top_commenters: list[dict],
    detail_comments: list[dict] | dict,
    summary_stats: dict,
    target_username: str,
    start_date: str,
    end_date: str,
    platform: str = "Instagram",
    filename: str | None = None,
) -> str:
    """Export data top commenters dan likes ke file Excel."""
    # Normalisasi detail_comments jika dikirim dalam bentuk dictionary
    comments_list: list[dict] = []
    if isinstance(detail_comments, dict):
        for user_comms in detail_comments.values():
            if isinstance(user_comms, list):
                comments_list.extend(user_comms)
            elif isinstance(user_comms, dict):
                comments_list.append(user_comms)
    elif isinstance(detail_comments, list):
        comments_list = detail_comments

    # Bersihkan nama file default jika belum ada
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        plat_slug = re.sub(r"[^\w\-]", "_", platform.lower()).strip("_")
        user_slug = re.sub(r"[^\w\-]", "_", target_username).strip("_")
        s_clean = start_date.replace("-", "")
        e_clean = end_date.replace("-", "")
        filename = f"top_commenters_{plat_slug}_{user_slug}_{s_clean}_{e_clean}_{timestamp}.xlsx"

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Info header
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = clean_cell_value(f"Top Commenters & Likes Analysis ({platform}) — @{target_username}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws_summary.merge_cells("A2:F2")
    ws_summary["A2"].value = clean_cell_value(f"Periode: {start_date} s/d {end_date}")
    ws_summary["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    # Statistik
    stats_start_row = 4
    stats_data = [
        ("Total Post di-scan", summary_stats.get("total_posts_scanned", 0)),
        ("Total Likes Postingan", summary_stats.get("total_post_likes", 0)),
        ("Rata-rata Likes/Post", summary_stats.get("avg_likes_per_post", 0)),
        ("Total Komentar", summary_stats.get("total_comments", 0)),
        ("Unique Commenters", summary_stats.get("unique_commenters", 0)),
        ("Rata-rata Komentar/Post", summary_stats.get("avg_comments_per_post", 0)),
    ]
    for i, (label, value) in enumerate(stats_data):
        row = stats_start_row + i
        ws_summary.cell(row=row, column=1, value=clean_cell_value(label)).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=clean_cell_value(value))

    # Tabel ranking top commenters
    table_start_row = stats_start_row + len(stats_data) + 2
    summary_headers = [
        "Rank",
        "Username",
        "Jumlah Komentar",
        "Komentar Pertama",
        "Sudah Like Post?",
        "Total Like Postingan",
        "Total Like Komentar",
        "Jumlah Post Dikomen",
        "Post URLs",
    ]

    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=table_start_row, column=col_idx, value=clean_cell_value(header))
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_fill_even = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for i, commenter in enumerate(top_commenters):
        row = table_start_row + 1 + i
        post_urls = commenter.get("post_urls", [])
        if isinstance(post_urls, list):
            urls_str = "\n".join(str(u) for u in post_urls)
        else:
            urls_str = str(post_urls or "")

        ws_summary.cell(row=row, column=1, value=commenter.get("rank", i + 1))
        ws_summary.cell(row=row, column=2, value=clean_cell_value(commenter.get("username", "unknown")))
        ws_summary.cell(row=row, column=3, value=commenter.get("comment_count", 0))
        ws_summary.cell(row=row, column=4, value=clean_cell_value(format_tanggal_indonesia(commenter.get("earliest_comment_date", "N/A"))))
        ws_summary.cell(row=row, column=5, value=clean_cell_value(commenter.get("has_liked_post", "N/A")))
        ws_summary.cell(row=row, column=6, value=commenter.get("total_post_likes", 0))
        ws_summary.cell(row=row, column=7, value=commenter.get("total_comment_likes", 0))
        ws_summary.cell(row=row, column=8, value=commenter.get("unique_posts_count", 0))
        
        url_cell = ws_summary.cell(row=row, column=9, value=clean_cell_value(urls_str))
        url_cell.alignment = Alignment(wrap_text=True)

        if i % 2 == 1:
            for col in range(1, 10):
                ws_summary.cell(row=row, column=col).fill = data_fill_even

    if platform.lower() == "tiktok":
        note_row = table_start_row + len(top_commenters) + 2
        ws_summary.cell(
            row=note_row,
            column=1,
            value="* Catatan TikTok: Data status like per user bersifat privat di platform TikTok (kolom bernilai N/A)."
        ).font = Font(name="Calibri", italic=True, size=9, color="7F7F7F")

    _auto_fit_columns(ws_summary)

    # ── Sheet 2: Detail Komentar ──────────────────────────────────────
    ws_detail = wb.create_sheet("Detail Komentar")

    detail_headers = [
        "Username",
        "Teks Komentar",
        "Sudah Like Post?",
        "Like Komentar",
        "Tanggal Komentar",
        "Post URL",
        "Like Postingan",
        "Tanggal Post",
        "Caption Post",
    ]
    _style_header(ws_detail, detail_headers, fill_color="2E75B6")

    for i, comment in enumerate(comments_list):
        row = i + 2
        ws_detail.cell(row=row, column=1, value=clean_cell_value(comment.get("commenter_username", "unknown")))
        ws_detail.cell(row=row, column=2, value=clean_cell_value(comment.get("comment_text", "")))
        ws_detail.cell(row=row, column=3, value=clean_cell_value(comment.get("has_liked_post", "N/A")))
        ws_detail.cell(row=row, column=4, value=comment.get("comment_likes", 0))
        ws_detail.cell(row=row, column=5, value=clean_cell_value(format_tanggal_indonesia(comment.get("comment_date", "N/A"))))
        ws_detail.cell(row=row, column=6, value=clean_cell_value(comment.get("post_url", "")))
        ws_detail.cell(row=row, column=7, value=comment.get("post_likes", 0))
        ws_detail.cell(row=row, column=8, value=clean_cell_value(format_tanggal_indonesia(comment.get("post_date", "N/A"))))
        ws_detail.cell(row=row, column=9, value=clean_cell_value(comment.get("post_caption", "")))

        if i % 2 == 1:
            for col in range(1, 10):
                ws_detail.cell(row=row, column=col).fill = data_fill_even

    _auto_fit_columns(ws_detail)

    # ── Sheet 3: Daftar Postingan ─────────────────────────────────────
    ws_posts = wb.create_sheet("Daftar Postingan")
    posts_headers = [
        "No",
        "Post URL",
        "Jumlah Likes",
        "Tanggal Post",
        "Caption Post",
    ]
    _style_header(ws_posts, posts_headers, fill_color="107C41")

    # Kumpulkan postingan unik dari comments_list
    unique_posts = []
    seen_urls = set()
    for c in comments_list:
        p_url = c.get("post_url")
        if p_url and p_url not in seen_urls:
            seen_urls.add(p_url)
            unique_posts.append({
                "post_url": p_url,
                "post_likes": c.get("post_likes", 0),
                "post_date": c.get("post_date", "N/A"),
                "post_caption": c.get("post_caption", ""),
            })

    unique_posts.sort(key=lambda x: x.get("post_likes", 0), reverse=True)

    for i, p in enumerate(unique_posts, 1):
        row = i + 1
        ws_posts.cell(row=row, column=1, value=i)
        ws_posts.cell(row=row, column=2, value=clean_cell_value(p.get("post_url", "")))
        ws_posts.cell(row=row, column=3, value=p.get("post_likes", 0))
        ws_posts.cell(row=row, column=4, value=clean_cell_value(format_tanggal_indonesia(p.get("post_date", "N/A"))))
        ws_posts.cell(row=row, column=5, value=clean_cell_value(p.get("post_caption", "")))

        if i % 2 == 0:
            for col in range(1, 6):
                ws_posts.cell(row=row, column=col).fill = data_fill_even

    _auto_fit_columns(ws_posts)

    # Freeze pane pada header
    ws_summary.freeze_panes = ws_summary.cell(row=table_start_row + 1, column=1)
    ws_detail.freeze_panes = "A2"
    ws_posts.freeze_panes = "A2"

    return _save_workbook_safely(wb, filename)


def export_links_to_excel(
    posts: list[dict],
    target_username: str,
    start_date: str,
    end_date: str,
    platform: str = "TikTok",
    filename: str | None = None,
) -> str:
    """Export daftar link postingan hasil scan ke file Excel."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        plat_slug = re.sub(r"[^\w\-]", "_", platform.lower()).strip("_")
        user_slug = re.sub(r"[^\w\-]", "_", target_username).strip("_")
        s_clean = start_date.replace("-", "")
        e_clean = end_date.replace("-", "")
        filename = f"links_{plat_slug}_{user_slug}_{s_clean}_{e_clean}_{timestamp}.xlsx"

    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Link Postingan"

    # Info header
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = clean_cell_value(f"Daftar Link Postingan ({platform}) — @{target_username}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"].value = clean_cell_value(f"Periode: {start_date} s/d {end_date} • Total: {len(posts)} Postingan")
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    table_start_row = 4
    headers = ["No", "Post URL", "Tanggal Post", "Tipe Post", "Jumlah Likes", "Caption Post"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=clean_cell_value(header))
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_fill_even = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    for i, p in enumerate(posts, 1):
        row = table_start_row + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=clean_cell_value(p.get("post_url", "")))
        ws.cell(row=row, column=3, value=clean_cell_value(format_tanggal_indonesia(p.get("post_date", "N/A"))))
        ws.cell(row=row, column=4, value=clean_cell_value(p.get("post_type", "Video").upper()))
        ws.cell(row=row, column=5, value=p.get("post_likes", 0))
        ws.cell(row=row, column=6, value=clean_cell_value(p.get("post_caption", "")))

        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = data_fill_even

    _auto_fit_columns(ws)
    ws.freeze_panes = ws.cell(row=table_start_row + 1, column=1)

    return _save_workbook_safely(wb, filename)
